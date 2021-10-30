import csv
import math
from re import I
import threading
from os import path
import os
from typing import Optional

from sglogging import SgLogSetup

from sgscrape.pause_resume import CrawlState
from sgscrape.sgrecord import SgRecord
from sgscrape.sgrecord_id import RecommendedRecordIds
from sgscrape.sgrecord_deduper import SgRecordDeduper

from openpyxl import load_workbook, Workbook
from util.util import Util

ONE_SHEET_LIMIT = 999999
SAVE_LIMIT = 100
ZORO_TYPE = "ZORO"
ZORO_RESULTS_BUCKET = 'zoro-results'

class SgWriter:
    """
    Write records to a csv file safely and efficiently.
    When provided with an `SgRecordDeduper`, it also deduplicates
    based on current, and past (i.e. resumed) records.

    Usage:
    ```
    def fetch_data(sgw: SgWriter):
        sgw.write_row(SgRecord(page_url="123.com/1", store_number=''))
        sgw.write_row(SgRecord(raw={'page_url':"123.com/2", 'store_number': ''}))
        sgw.write_row(SgRecord(raw={'page_url':"123.com/3", 'store_number': ''}))

    with SgWriter(deduper=SgRecordDeduper(RecommendedRecordIds.PageUrlId)) as writer:
        fetch_data(writer)
    ```
    """

    __lock = threading.Lock()

    def __init__(self,
                 deduper = None,
                 data_file = CrawlState.DEFAULT_DATA_FILE,
                 s3 = None,
                 type = ZORO_TYPE):
        """
        Creates the writer. Note that the constructor also writes the header row (if necessary).

        :param deduper: Optionally, inject a record deduplicator to filter out duplicates. If present, the data file will
                        contain an additional colon-delimited row that represents the record's unique id creation.
        """
        self.__log = SgLogSetup().get_logger(logger_name='sgwriter')
        self.__deduper = deduper
        self.__s3 = s3
        self.__data_file_name = data_file
        preexisting = path.exists(data_file)
        if preexisting:
            os.remove(data_file)
            preexisting = False
        self.__zoro_wb = load_workbook(data_file) if preexisting else Workbook()
        self.__writer = self.__zoro_wb.active
        self.__id_str = str(deduper.get_id()) if deduper else None
        # header_row = SgRecord.Headers.HEADER_ROW_WITH_REC_ID if self.__id_str else SgRecord.Headers.HEADER_ROW
        self.header_row = SgRecord.Headers.ZORO_ROW if type == ZORO_TYPE else SgRecord.Headers.BH_ROW
        if not preexisting:
            self.__write_header()
        else:
            self.__log.debug(f'Pre-existing data file found: {data_file} . Skipping header row.')

    def __enter__(self):
        return self

    def save_file(self):
        self.__log.info('save and upload file to s3')
        self.__zoro_wb.save(self.__data_file_name)
        self.__s3.upload_file(self.__data_file_name, ZORO_RESULTS_BUCKET, self.__data_file_name.split('/')[-1])

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save_file()

    def __write_header(self)-> Optional[str]:
        SgWriter.__lock.acquire()
        for col, val in enumerate(self.header_row):
            self.__writer.cell(1, col+1).value = val
        SgWriter.__lock.release()

    def write_row(self, record: SgRecord) -> Optional[str]:
        SgWriter.__lock.acquire()
        max_row = Util().max_row(self.__writer)
        if max_row % SAVE_LIMIT == 0 and max_row > 2:
            self.save_file()

        # if one sheet exceeds the limit, then create new sheet and write header properly
        if max_row > ONE_SHEET_LIMIT:
            self.__log.info('create new sheet')
            next_row = 2
            self.__writer = self.__zoro_wb.create_sheet()
            SgWriter.__lock.release()
            self.__write_header()
        else:
            SgWriter.__lock.release()
            next_row = max_row + 1
       

        is_dup, id = self.__deduper.dedup(record)
        if not is_dup:
            SgWriter.__lock.acquire()
            for column, val in record.as_dict().items():
                col = Util().get_col_from_name(self.__writer, column)
                self.__writer.cell(next_row, col).value = val
            SgWriter.__lock.release()
            return None
        else:
            return id

    # def write_row(self, record: SgRecord) -> Optional[str]:
    #     """
    #     Writes the record to file, or else returns the id of a record, only if it is a duplicate.
    #     """
    #     if not self.__deduper:
    #         SgWriter.__lock.acquire()
    #         self.__writer.writerow(record.as_row())
    #         SgWriter.__lock.release()
    #         return None
    #     else:
    #         is_dup, id = self.__deduper.dedup(record)
    #         if not is_dup:
    #             row = record.as_row()
    #             row.append(self.__id_str)
    #             SgWriter.__lock.acquire()
    #             self.__writer.writerow(row)
    #             SgWriter.__lock.release()
    #             return None
    #         else:
    #             return id


if __name__ == "__main__":
    def fetch_data(sgw: SgWriter):
        sgw.write_row(SgRecord(page_url="123.com/1", store_number='1', latitude=math.pi, longitude=math.e))
        sgw.write_row(SgRecord(raw={'page_url':"123.com/2", 'store_number': 2}))
        sgw.write_row(SgRecord(raw={'page_url':"123.com/3", 'store_number': math.pi}))

    with SgWriter(deduper=SgRecordDeduper(RecommendedRecordIds.PageUrlId)) as writer:
        fetch_data(writer)

    # writes page_url as a string, and everything else, including `store_data`, as <MISSING>.